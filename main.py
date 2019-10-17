import time
import xlwt
import sys
import os


def Uexit(cacheFile):
    if cacheFile:
        try:
            os.remove(cacheFile)
        except:
            input("\n回车键关闭")
            sys.exit(1)
    input("\n回车键关闭")
    sys.exit(1)


NowTime = time.time()

# if NowTime > 1577808000000:
# 	print("错误02")
# 	Uexit(False)

print("欢迎使用表格合并，直接运行软件查看帮助，拖动表格文件至软件进行合并处理。\n")
if len(sys.argv) == 1:
    print("0: github.com/queziaa/Excel_merging_two_tables")
    print("1: 输入文件需要拥有两个工作簿，名字任意")
    print("2：将以第一列的内容进行比对将两表合并")
    print("3：同一工作簿第一列不能出现相同内容")
    print("4：第一行视为表头直接复制至结果")
    print("5: 产生遗弃内容时，遗弃内容会在原工作簿同名工作簿下保存")
    Uexit(False)


def shenfenzhengFormat(temp):
    return str(temp).replace(" ", "").split(".")[0]


def isRepeat(temp):
    dic = {}
    index = 0
    for i in temp:
        a = shenfenzhengFormat(i)
        if a in dic:
            return [True, a, str(dic[a]+1), str(index+1)]
        dic[a] = index
        index += 1
    return [False]


def sheetAdd(sleet, index, list_1, list_2):
    colu = 0
    for i in list_1 + list_2[1:]:
        sleet.write(index, colu, i)
        colu += 1


def openpyxl_list(sleet, rows):
    if xls:
        return sleet.row_values(rows)
    temp = []
    for cell in list(sleet.rows)[rows]:
        temp.append(cell.value)
    return temp


oldBookUrl = os.path.split(sys.argv[1])
xls = oldBookUrl[1].split(".")[-1]
if xls == 'xlsx':
    xls = False
    a = "xlsx"
elif xls == "xls":
    xls = True
    a = "xls"
else:
    print("输入文件不是表格")
    Uexit(False)
cacheFile = oldBookUrl[0] + '\\缓存' + \
    str(NowTime).split(".")[1] + str(NowTime).split(".")[0] + '.' + a
a = 'copy ' + '"' + sys.argv[1] + '" ' + '"' + cacheFile + '"'
os.popen(a)

mode = input(
    "请输入工作模式\n 1: 生成一表和二表共有部分 （一交二）\n 2: 生成一表全部 （一）\n 3: 生成一表和二表全部 （一并二）\n")
if mode != '1' and mode != '2' and mode != '3':
    print("输入错误")
    Uexit(cacheFile)

i = 0
while True:
    i += 1
    if os.path.exists(cacheFile):
        break
    time.sleep(0.1)
    if i > 100:
        print("错误01")
        Uexit(cacheFile)

if xls:
    import xlrd
    workbook = xlrd.open_workbook(cacheFile)
    shenames = workbook.sheet_names()
    if len(shenames) == 1:
        print("输入表格需要拥有两个工作簿")
    work_1 = workbook.sheet_by_index(0)
    work_2 = workbook.sheet_by_index(1)
    work_1_rows = work_1.nrows
    work_1_columns = work_1.ncols
    work_2_rows = work_2.nrows
    work_2_columns = work_2.ncols
else:
    import openpyxl
    workbook = openpyxl.load_workbook(cacheFile)
    shenames = workbook.sheetnames
    if len(shenames) == 1:
        print("输入表格需要拥有两个工作簿")
    work_1 = workbook[shenames[0]]
    work_2 = workbook[shenames[1]]
    work_1_rows = work_1.max_row
    work_1_columns = work_1.max_column
    work_2_rows = work_2.max_row
    work_2_columns = work_2.max_column

print("表一   行数 " + str(work_1_rows) + " 列数" + str(work_1_columns))
print("表二   行数 " + str(work_2_rows) + " 列数" + str(work_2_columns) + '\n')

work_2_observe = [x for x in range(work_2_rows)[1:]]
newBook = xlwt.Workbook(encoding="utf-8", style_compression=0)
sheet_3 = newBook.add_sheet("处理结果", cell_overwrite_ok=True)
temp = openpyxl_list(work_1, 0) + openpyxl_list(work_2, 0)[1:]
sheet_3Column = 0
for i in temp:
    sheet_3.write(0, sheet_3Column, i)
    sheet_3Column += 1
if xls:
    work_1_Rep = work_1.col_values(0)
    work_2_Rep = work_2.col_values(0)
else:
    work_1_Rep = []
    work_2_Rep = []
    for cell in list(work_1.columns)[0]:
        work_1_Rep.append(cell.value)
    for cell in list(work_2.columns)[0]:
        work_2_Rep.append(cell.value)
work_1_Rep = isRepeat(work_1_Rep)
if work_1_Rep[0]:
    print("表一第一列有重复 重复项为" + work_1_Rep[1] + "同时出现在" +
          work_1_Rep[2] + "行与" + work_1_Rep[3] + "行")
    Uexit(cacheFile)
work_2_Rep = isRepeat(work_2_Rep)
if work_2_Rep[0]:
    print("表二第一列有重复 重复项为" + work_2_Rep[1] + "同时出现在" +
          work_2_Rep[2] + "行与" + work_2_Rep[3] + "行\n")
    Uexit(cacheFile)

index = 1
sheet_1_index = 1
sheet_1 = False
for work_1_i in range(work_1_rows)[1:]:
    print("已经处理 "+str(work_1_i/work_1_rows*100)[:4]+"% 的数据")
    work_1_list = openpyxl_list(work_1, work_1_i)
    work_1_mark = shenfenzhengFormat(work_1_list[0])
    found = False
    for work_2_i in work_2_observe:
        work_2_list = openpyxl_list(work_2, work_2_i)
        if work_1_mark == shenfenzhengFormat(work_2_list[0]):
            sheetAdd(sheet_3, index, work_1_list, work_2_list)
            index += 1
            found = True
            work_2_observe.remove(work_2_i)
            break
    if not found:
        if mode == '1':
            if sheet_1_index == 1:
                sheet_1 = newBook.add_sheet(
                    shenames[0], cell_overwrite_ok=True)
            sheetAdd(sheet_1, sheet_1_index, work_1_list, [])
            sheet_1_index += 1
        else:
            sheetAdd(sheet_3, index, work_1_list, [])
            index += 1

if mode == '1' and sheet_1_index != 1:
    temp = openpyxl_list(work_1, 0)
    sheetAdd(sheet_1, 0, temp, [])

if len(work_2_observe) != 0:
    if mode == '1' or mode == '2':
        sheet_2 = newBook.add_sheet(shenames[1], cell_overwrite_ok=True)
        temp = []
        temp = openpyxl_list(work_2, 0)
        sheetAdd(sheet_2, 0, temp, [])
        index = 1
        for i in work_2_observe:
            work_2_list = openpyxl_list(work_2, i)
            sheetAdd(sheet_2, index, work_2_list, [])
            index += 1
    else:
        for i in work_2_observe:
            work_2_list = openpyxl_list(work_2, i)
            ii = ['' for x in range(work_1_columns)]
            ii[0] = work_2_list[0]
            sheetAdd(sheet_3, index, ii, work_2_list)
            index += 1

newfil = cacheFile.replace("缓存", "结果").replace("xlsx", "xls")
newBook.save(newfil)
print("\n处理完成 文件名为：       " + newfil)
Uexit(cacheFile)
